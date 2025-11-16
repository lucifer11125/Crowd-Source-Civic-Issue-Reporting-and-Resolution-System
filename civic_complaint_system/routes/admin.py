import csv
import io
from datetime import datetime, timedelta
from flask import render_template, request, redirect, url_for, flash, make_response, current_app
from flask_login import login_required, current_user
from models import db, Complaint, User, StatusUpdate
from routes import admin_bp
from routes.auth import role_required
from sqlalchemy import func, and_, or_, case
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def validate_user_form(data, user_id=None):
    """Validate user form data"""
    errors = []

    name = data.get('name', '').strip()
    email = data.get('email', '').strip().lower()
    role = data.get('role')
    department = data.get('department', '').strip()
    is_active = data.get('is_active') == 'on'

    if not name or len(name) < 2:
        errors.append('Name must be at least 2 characters long')

    if not email:
        errors.append('Email is required')
    else:
        # Check for duplicate email (excluding current user)
        query = User.query.filter_by(email=email)
        if user_id:
            query = query.filter(User.id != user_id)
        if query.first():
            errors.append('Email already exists')

    if role not in ['citizen', 'municipal', 'admin']:
        errors.append('Invalid role selected')

    if role == 'municipal' and not department:
        errors.append('Department is required for municipal officers')

    return errors

@admin_bp.route('/admin/dashboard')
@login_required
@role_required('admin')
def admin_dashboard():
    """Admin overview with statistics"""
    # User statistics
    total_users = User.query.count()
    citizens_count = User.query.filter_by(role='citizen').count()
    officers_count = User.query.filter_by(role='municipal').count()
    admins_count = User.query.filter_by(role='admin').count()

    # Complaint statistics
    total_complaints = Complaint.query.count()
    submitted_count = Complaint.query.filter_by(status='submitted').count()
    in_progress_count = Complaint.query.filter_by(status='in_progress').count()
    resolved_count = Complaint.query.filter_by(status='resolved').count()
    rejected_count = Complaint.query.filter_by(status='rejected').count()

    # Department breakdown
    department_stats = db.session.query(
        Complaint.assigned_department,
        func.count(Complaint.id).label('complaint_count')
    ).filter(Complaint.assigned_department.isnot(None))\
     .group_by(Complaint.assigned_department)\
     .all()

    # Recent activity (last 7 days)
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    recent_complaints = Complaint.query.filter(Complaint.created_at >= seven_days_ago).count()
    recent_resolutions = Complaint.query.filter(
        Complaint.resolved_at >= seven_days_ago
    ).count()

    # Complaint trends by category
    category_trends = db.session.query(
        Complaint.category,
        func.count(Complaint.id).label('total'),
        func.sum(case((Complaint.status == 'resolved', 1), else_=0)).label('resolved')
    ).group_by(Complaint.category).all()

    # Trend analysis: Daily submissions and resolutions for last 7 days
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    submission_trends = db.session.query(
        func.date(Complaint.created_at).label('date'),
        func.count(Complaint.id).label('count')
    ).filter(Complaint.created_at >= seven_days_ago)\
     .group_by(func.date(Complaint.created_at))\
     .order_by(func.date(Complaint.created_at)).all()

    # Convert Row objects to dictionaries for JSON serialization
    submission_trends = [{'date': str(row.date), 'count': row.count} for row in submission_trends]

    resolution_trends = db.session.query(
        func.date(Complaint.resolved_at).label('date'),
        func.count(Complaint.id).label('count')
    ).filter(Complaint.resolved_at >= seven_days_ago)\
     .group_by(func.date(Complaint.resolved_at))\
     .order_by(func.date(Complaint.resolved_at)).all()

    # Convert Row objects to dictionaries for JSON serialization
    resolution_trends = [{'date': str(row.date), 'count': row.count} for row in resolution_trends]

    # Rejected trends
    rejected_trends = db.session.query(
        func.date(Complaint.updated_at).label('date'),
        func.count(Complaint.id).label('count')
    ).filter(Complaint.status == 'rejected')\
     .filter(Complaint.updated_at >= seven_days_ago)\
     .group_by(func.date(Complaint.updated_at))\
     .order_by(func.date(Complaint.updated_at)).all()

    rejected_trends = [{'date': str(row.date), 'count': row.count} for row in rejected_trends]

    # In progress trends
    in_progress_trends = db.session.query(
        func.date(Complaint.updated_at).label('date'),
        func.count(Complaint.id).label('count')
    ).filter(Complaint.status == 'in_progress')\
     .filter(Complaint.updated_at >= seven_days_ago)\
     .group_by(func.date(Complaint.updated_at))\
     .order_by(func.date(Complaint.updated_at)).all()

    in_progress_trends = [{'date': str(row.date), 'count': row.count} for row in in_progress_trends]

    # Unassigned complaints
    unassigned_count = Complaint.query.filter_by(assigned_officer=None)\
        .filter(Complaint.status.in_(['submitted', 'in_progress'])).count()

    return render_template('admin_dashboard.html',
                         # User stats
                         total_users=total_users,
                         citizens_count=citizens_count,
                         officers_count=officers_count,
                         admins_count=admins_count,
                         # Complaint stats
                         total_complaints=total_complaints,
                         submitted_count=submitted_count,
                         in_progress_count=in_progress_count,
                         resolved_count=resolved_count,
                         rejected_count=rejected_count,
                         # Department stats
                         department_stats=department_stats,
                         # Recent activity
                         recent_complaints=recent_complaints,
                         recent_resolutions=recent_resolutions,
                         # Trends
                         category_trends=category_trends,
                         submission_trends=submission_trends,
                         resolution_trends=resolution_trends,
                         rejected_trends=rejected_trends,
                         in_progress_trends=in_progress_trends,
                         unassigned_count=unassigned_count)

@admin_bp.route('/admin/users')
@login_required
@role_required('admin')
def users():
    """User management page"""
    role_filter = request.args.get('role', 'all')
    status_filter = request.args.get('status', 'all')

    query = User.query

    # Apply filters
    if role_filter != 'all':
        query = query.filter_by(role=role_filter)

    if status_filter == 'active':
        query = query.filter_by(is_active=True)
    elif status_filter == 'inactive':
        query = query.filter_by(is_active=False)

    users = query.order_by(User.created_at.desc()).all()

    return render_template('users.html',
                         users=users,
                         role_filter=role_filter,
                         status_filter=status_filter)

@admin_bp.route('/admin/users/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def edit_user(id):
    """Edit user details"""
    user = User.query.get_or_404(id)

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip().lower()
        role = request.form.get('role')
        department = request.form.get('department', '').strip()
        is_active = request.form.get('is_active') == 'on'

        # Validate form
        errors = validate_user_form({
            'name': name,
            'email': email,
            'role': role,
            'department': department,
            'is_active': is_active
        }, user.id)

        if errors:
            for error in errors:
                flash(error, 'danger')
            return render_template('edit_user.html', user=user)

        # Update user
        user.name = name
        user.email = email
        user.role = role
        user.department = department if role == 'municipal' or role == 'admin' else None
        user.is_active = is_active

        try:
            db.session.commit()
            flash(f'User {user.name} updated successfully!', 'success')
            return redirect(url_for('admin.users'))
        except Exception as e:
            db.session.rollback()
            flash('Failed to update user. Please try again.', 'danger')

    return render_template('edit_user.html', user=user)

@admin_bp.route('/admin/users/<int:id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def delete_user(id):
    """Delete user (with confirmation)"""
    user = User.query.get_or_404(id)

    # Prevent deletion of users with active complaints
    active_complaints = Complaint.query.filter(
        or_(
            Complaint.user_id == user.id,
            Complaint.assigned_officer == user.id
        ),
        Complaint.status.in_(['submitted', 'in_progress'])
    ).count()

    if active_complaints > 0:
        flash(f'Cannot delete user {user.name}. They have {active_complaints} active complaints.', 'danger')
        return redirect(url_for('admin.users'))

    # Prevent self-deletion
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('admin.users'))

    try:
        db.session.delete(user)
        db.session.commit()
        flash(f'User {user.name} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Failed to delete user. Please try again.', 'danger')

    return redirect(url_for('admin.users'))

@admin_bp.route('/admin/reports')
@login_required
@role_required('admin')
def reports():
    """Generate and download reports"""
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status_filter = request.args.get('status', 'all')
    category_filter = request.args.get('category', 'all')
    department_filter = request.args.get('department', 'all')

    # Parse dates
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            start_date = None
            flash('Invalid start date format', 'danger')
    else:
        start_date = datetime.utcnow() - timedelta(days=30)  # Default to last 30 days

    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        except ValueError:
            end_date = None
            flash('Invalid end date format', 'danger')
    else:
        end_date = datetime.utcnow()

    # Build query
    query = Complaint.query.filter(
        Complaint.created_at >= start_date,
        Complaint.created_at <= end_date
    )

    # Apply filters
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)

    if category_filter != 'all':
        query = query.filter_by(category=category_filter)

    if department_filter != 'all':
        query = query.filter(Complaint.assigned_department == department_filter)

    complaints = query.order_by(Complaint.created_at.desc()).all()

    # Get available options for filters
    departments = db.session.query(User.department)\
                           .filter(User.department.isnot(None))\
                           .distinct().all()
    departments = [dept[0] for dept in departments]

    # Check if CSV export is requested
    if request.args.get('export') == 'csv':
        return generate_csv_report(complaints)

    # Check if Excel export is requested
    if request.args.get('export') == 'excel':
        return generate_excel_report(complaints)

    return render_template('reports.html',
                         complaints=complaints,
                         start_date=start_date.strftime('%Y-%m-%d') if start_date else '',
                         end_date=end_date.strftime('%Y-%m-%d') if end_date else '',
                         status_filter=status_filter,
                         category_filter=category_filter,
                         department_filter=department_filter,
                         departments=departments)

def generate_csv_report(complaints):
    """Generate CSV report for complaints"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow([
        'Complaint ID',
        'Submission Date',
        'Citizen Name',
        'Citizen Email',
        'Category',
        'Priority',
        'Address',
        'Landmark',
        'Description',
        'Status',
        'Assigned Officer',
        'Department',
        'Resolution Notes',
        'Created At',
        'Updated At',
        'Resolved At',
        'Status Updates Count'
    ])

    # Write data rows
    for complaint in complaints:
        # Get status updates count
        updates_count = StatusUpdate.query.filter_by(complaint_id=complaint.id).count()

        writer.writerow([
            complaint.id,
            complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            complaint.user.name,
            complaint.user.email,
            complaint.category,
            complaint.priority,
            complaint.address,
            complaint.landmark or '',
            complaint.description,
            complaint.status,
            complaint.assigned_officer.name if complaint.assigned_officer else 'Unassigned',
            complaint.assigned_officer.department if complaint.assigned_officer else '',
            complaint.resolution_notes or '',
            complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            complaint.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            complaint.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if complaint.resolved_at else '',
            updates_count
        ])

    # Create response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=complaints_report_{datetime.now().strftime("%Y-%m-%d")}.csv'

    return response

def generate_excel_report(complaints):
    """Generate Excel report for complaints"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Complaints Report"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center")

    # Write header
    headers = [
        'Complaint ID',
        'Submission Date',
        'Citizen Name',
        'Citizen Email',
        'Category',
        'Priority',
        'Address',
        'Landmark',
        'Description',
        'Status',
        'Assigned Officer',
        'Department',
        'Resolution Notes',
        'Created At',
        'Updated At',
        'Resolved At',
        'Status Updates Count'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Write data rows
    for row_num, complaint in enumerate(complaints, 2):
        # Get status updates count
        updates_count = StatusUpdate.query.filter_by(complaint_id=complaint.id).count()

        data = [
            complaint.id,
            complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            complaint.user.name,
            complaint.user.email,
            complaint.category,
            complaint.priority,
            complaint.address,
            complaint.landmark or '',
            complaint.description,
            complaint.status,
            complaint.assigned_officer.name if complaint.assigned_officer else 'Unassigned',
            complaint.assigned_officer.department if complaint.assigned_officer else '',
            complaint.resolution_notes or '',
            complaint.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            complaint.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            complaint.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if complaint.resolved_at else '',
            updates_count
        ]

        for col_num, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=complaints_report_{datetime.now().strftime("%Y-%m-%d")}.xlsx'

    return response

@admin_bp.route('/admin/complaints')
@login_required
@role_required('admin')
def all_complaints():
    """View all complaints (admin only)"""
    status_filter = request.args.get('status', 'all')
    category_filter = request.args.get('category', 'all')
    department_filter = request.args.get('department', 'all')

    # Build query
    query = Complaint.query

    # Apply filters
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)

    if category_filter != 'all':
        query = query.filter_by(category=category_filter)

    if department_filter != 'all':
        query = query.filter(Complaint.assigned_department == department_filter)

    complaints = query.order_by(Complaint.created_at.desc()).all()

    # Get available options for filters
    departments = db.session.query(User.department)\
                           .filter(User.department.isnot(None))\
                           .distinct().all()
    departments = [dept[0] for dept in departments]

    return render_template('all_complaints.html',
                         complaints=complaints,
                         status_filter=status_filter,
                         category_filter=category_filter,
                         department_filter=department_filter,
                         departments=departments)